
from pathlib import Path
from datetime import datetime
import shutil
import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

APP = "Controle Operacional Local / Online"
DATA_DIR = Path("dados")
BACKUP_DIR = DATA_DIR / "backups"
MASTER_FILE = DATA_DIR / "excel_mestre_operacional.xlsx"

DATA_DIR.mkdir(exist_ok=True)
BACKUP_DIR.mkdir(exist_ok=True)

COLS = [
    "Data", "Motorista", "Documento", "Cliente", "Unidade", "Paletes", "Valor",
    "L", "C", "F", "Status", "Tipo", "CPF", "Cavalo", "Carreta",
    "CS_OK", "C_OK", "L_OK", "F_OK", "Observações", "Inconsistências"
]

FILL_HEADER = "1F4E78"
FILL_L = "BDD7EE"
FILL_C = "FFF2CC"
FILL_F = "C6E0B4"
FILL_PROBLEM = "FFC7CE"
FILL_MONEY = "E2F0D9"

MOTORISTAS = {
    "GABRIEL BORGES": {"CPF": "809.066.155-87", "Cavalo": "KIY3204", "Carreta": "KGG1152"},
    "VALDEMIR DE JESUS": {"CPF": "044.327.095-37", "Cavalo": "KFL0115", "Carreta": ""},
    "JEAN ROBSON": {"CPF": "032.795.865-00", "Cavalo": "HWB9F22", "Carreta": ""},
    "WILSON REIS": {"CPF": "806.984.765-49", "Cavalo": "JJF1856", "Carreta": "NKZ6545"},
    "FABIO SOUZA": {"CPF": "007.714.835-30", "Cavalo": "PEJ4695", "Carreta": "NVQ8447"},
    "LUIS CARLOS": {"CPF": "934.560.345-04", "Cavalo": "KLB5018", "Carreta": "DTD8506"},
    "ARGEMIRO BORGES": {"CPF": "041.604.865-09", "Cavalo": "PEG7666", "Carreta": ""},
    "JONES ROSARIO": {"CPF": "533.594.654-00", "Cavalo": "JHX3C33", "Carreta": "KKT9007"},
}

def norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ["none", "nan", "nat"] else s

def norm_upper(v):
    return norm(v).upper()

def only_digits(v):
    return "".join(ch for ch in norm(v) if ch.isdigit())

def day_sheet_name(data):
    return norm(data).replace("/", "-") if norm(data) else "Sem data"

def create_empty_workbook(path=MASTER_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mestre"
    ws.append(COLS)
    add_rules_sheet(wb)
    style_all(wb)
    wb.save(path)

def add_rules_sheet(wb):
    if "Regras" in wb.sheetnames:
        del wb["Regras"]
    ws = wb.create_sheet("Regras")
    rows = [
        ["Regra", "Descrição"],
        ["Sem custo de API", "Este sistema guarda, organiza, busca, edita e remove registros do Excel mestre."],
        ["Importação", "Importe o Excel atualizado do ChatGPT para consolidar histórico."],
        ["Edição", "Pesquise uma delivery e edite/remova o registro na aba Buscar / Editar."],
        ["Histórico", "Nunca apagar histórico sem confirmação manual."],
        ["L/C/F", "Usar somente horários confirmados conforme interpretação operacional."],
        ["Sem coleta", "Quando houver deslocamento, D ou sem coleta, não criar C."],
        ["Remessa", "Remessa 3401 pode ter apenas CS/C/L/F OK, sem horários."],
        ["Bonocô Jean", "Delivery Bonocô do Jean termina com 3871."],
        ["JDE CAFÉ", "Remessas JDE CAFÉ devem manter exatamente esse nome."],
        ["Siglas", "F.S = Feira de Santana; L.F = Lauro de Freitas; S.F = Simões Filho."],
    ]
    for r in rows:
        ws.append(r)

def style_ws(ws):
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
            header = headers[cell.column-1] if cell.column-1 < len(headers) else ""
            if header == "L":
                cell.fill = PatternFill("solid", fgColor=FILL_L)
            elif header == "C":
                cell.fill = PatternFill("solid", fgColor=FILL_C)
            elif header == "F":
                cell.fill = PatternFill("solid", fgColor=FILL_F)
            elif header == "Valor":
                cell.fill = PatternFill("solid", fgColor=FILL_MONEY)
            elif header == "Inconsistências" and cell.value:
                cell.fill = PatternFill("solid", fgColor=FILL_PROBLEM)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 12, "B": 22, "C": 18, "D": 34, "E": 18, "F": 10, "G": 12,
        "H": 10, "I": 10, "J": 10, "K": 22, "L": 16, "M": 16, "N": 12,
        "O": 12, "P": 9, "Q": 9, "R": 9, "S": 9, "T": 45, "U": 45
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def style_all(wb):
    for ws in wb.worksheets:
        style_ws(ws)

def backup_current():
    if MASTER_FILE.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy2(MASTER_FILE, BACKUP_DIR / f"backup_excel_mestre_{ts}.xlsx")

def ensure_workbook():
    if not MASTER_FILE.exists():
        create_empty_workbook()

def ensure_day_sheet(wb, data):
    name = day_sheet_name(data)
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(COLS)
    return wb[name]

def get_headers(ws):
    return [c.value for c in ws[1]]

def find_record_rows(ws, documento=None, data=None, motorista=None, cliente=None):
    doc = only_digits(documento)
    matches = []
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_doc = only_digits(row[2].value)
        row_data = norm(row[0].value)
        row_motorista = norm_upper(row[1].value)
        row_cliente = norm_upper(row[3].value)
        ok = True
        if doc and row_doc != doc:
            ok = False
        if data and row_data != norm(data):
            ok = False
        if motorista and norm_upper(motorista) not in row_motorista:
            ok = False
        if cliente and norm_upper(cliente) not in row_cliente:
            ok = False
        if ok:
            matches.append(idx)
    return matches

def find_record_row(ws, documento, data=None):
    rows = find_record_rows(ws, documento=documento, data=data)
    if rows:
        return rows[0]
    rows = find_record_rows(ws, documento=documento)
    return rows[0] if rows else None

def enrich(row):
    motorista = norm_upper(row.get("Motorista"))
    for name, data in MOTORISTAS.items():
        if motorista and (motorista in name or name in motorista):
            row["Motorista"] = name.title()
            row["CPF"] = data["CPF"]
            row["Cavalo"] = data["Cavalo"]
            row["Carreta"] = data["Carreta"]
            break

    doc = only_digits(row.get("Documento"))
    cliente = norm_upper(row.get("Cliente"))
    unidade = norm_upper(row.get("Unidade"))
    mot = norm_upper(row.get("Motorista"))

    if "JEAN" in mot and ("BONOCO" in cliente or "BONOCÔ" in cliente or "BONOCO" in unidade or "BONOCÔ" in unidade):
        if doc.endswith("3816"):
            row["Documento"] = doc[:-4] + "3871"
            row["Observações"] = (norm(row.get("Observações")) + " | Corrigido Bonocô Jean final 3871").strip(" |")

    if doc.startswith("3401") and ("JDE" in cliente or "CAFÉ" in cliente or "CAFE" in cliente):
        row["Cliente"] = "JDE CAFÉ"

    status_context = " ".join([norm_upper(row.get("Status")), norm_upper(row.get("Tipo")), norm_upper(row.get("Observações"))])
    if any(x in status_context for x in ["DESLOCAMENTO", "SEM COLETA", "BLOQUEIO"]):
        row["C"] = ""
        if not norm(row.get("Tipo")):
            row["Tipo"] = "deslocamento"
    return row

def update_ws(ws, row, overwrite=False):
    row = enrich(row.copy())
    data = norm(row.get("Data"))
    doc = norm(row.get("Documento"))
    idx = find_record_row(ws, doc, data)

    values = [row.get(c, "") for c in COLS]

    if idx is None:
        ws.append(values)
        return "criado"

    headers = get_headers(ws)
    for col_idx, col_name in enumerate(COLS, start=1):
        new_val = row.get(col_name, "")
        cell = ws.cell(idx, col_idx)
        if overwrite:
            cell.value = new_val
            continue
        if new_val in ["", None, False]:
            continue
        if col_name in ["L", "C", "F"] and cell.value and norm(cell.value) != norm(new_val):
            obs_col = headers.index("Observações") + 1
            old_obs = norm(ws.cell(idx, obs_col).value)
            ws.cell(idx, obs_col).value = (old_obs + f" | {col_name} mantido {cell.value}; novo sugerido {new_val}").strip(" |")
            continue
        cell.value = new_val
    return "atualizado"

def delete_record_from_ws(ws, documento, data=None, motorista=None, cliente=None, delete_incomplete_only=False):
    rows = find_record_rows(ws, documento=documento, data=data, motorista=motorista, cliente=cliente)
    deleted = 0
    for idx in sorted(rows, reverse=True):
        vals = [ws.cell(idx, c).value for c in range(1, len(COLS)+1)]
        if delete_incomplete_only:
            filled = sum(1 for v in vals if norm(v))
            # apaga só linha quase vazia, como None/None/None/Documento perdido em Cliente
            if filled > 3:
                continue
        ws.delete_rows(idx, 1)
        deleted += 1
    return deleted

def workbook_to_df():
    ensure_workbook()
    wb = load_workbook(MASTER_FILE, data_only=True)
    ws = wb["Mestre"]
    data = list(ws.values)
    if len(data) <= 1:
        return pd.DataFrame(columns=COLS)
    df = pd.DataFrame(data[1:], columns=data[0]).dropna(how="all")
    for c in COLS:
        if c not in df.columns:
            df[c] = ""
    return df[COLS]

def import_excel(uploaded):
    tmp = DATA_DIR / "importado.xlsx"
    tmp.write_bytes(uploaded.getvalue())

    backup_current()
    ensure_workbook()

    source = load_workbook(tmp, data_only=True)
    target = load_workbook(MASTER_FILE)

    s_ws = source["Mestre"] if "Mestre" in source.sheetnames else source[source.sheetnames[0]]
    rows = list(s_ws.values)
    if not rows:
        return 0

    headers = [norm(h) for h in rows[0]]
    count = 0

    for raw in rows[1:]:
        if not any(raw):
            continue
        row = {}
        for i, h in enumerate(headers):
            if h in COLS:
                row[h] = raw[i] if i < len(raw) else ""
        if not row.get("Documento") and len(raw) >= 3:
            row["Documento"] = raw[2]
        if not row.get("Data") and len(raw) >= 1:
            row["Data"] = raw[0]
        if not row.get("Motorista") and len(raw) >= 2:
            row["Motorista"] = raw[1]

        # evita importar linhas vazias/erradas
        if not only_digits(row.get("Documento")):
            continue

        m_ws = target["Mestre"]
        d_ws = ensure_day_sheet(target, row.get("Data"))
        update_ws(m_ws, row)
        update_ws(d_ws, row)
        count += 1

    add_rules_sheet(target)
    style_all(target)
    target.save(MASTER_FILE)
    return count

def add_manual_record(row):
    # trava contra registros manuais incompletos
    if not only_digits(row.get("Documento")):
        raise ValueError("Documento/delivery é obrigatório.")
    if not norm(row.get("Data")):
        raise ValueError("Data é obrigatória.")
    if not norm(row.get("Motorista")) and not norm(row.get("Cliente")):
        raise ValueError("Informe pelo menos motorista ou cliente.")

    backup_current()
    ensure_workbook()
    wb = load_workbook(MASTER_FILE)
    update_ws(wb["Mestre"], row)
    update_ws(ensure_day_sheet(wb, row.get("Data")), row)
    add_rules_sheet(wb)
    style_all(wb)
    wb.save(MASTER_FILE)

def edit_record(original_doc, original_data, row):
    backup_current()
    ensure_workbook()
    wb = load_workbook(MASTER_FILE)

    # remove o registro antigo de Mestre e aba do dia original
    delete_record_from_ws(wb["Mestre"], original_doc, data=original_data)
    old_sheet = day_sheet_name(original_data)
    if old_sheet in wb.sheetnames:
        delete_record_from_ws(wb[old_sheet], original_doc, data=original_data)

    # grava registro novo/editado
    update_ws(wb["Mestre"], row, overwrite=True)
    update_ws(ensure_day_sheet(wb, row.get("Data")), row, overwrite=True)

    add_rules_sheet(wb)
    style_all(wb)
    wb.save(MASTER_FILE)

def delete_record(documento, data=None, motorista=None, cliente=None, incomplete_only=False):
    backup_current()
    ensure_workbook()
    wb = load_workbook(MASTER_FILE)
    total = delete_record_from_ws(wb["Mestre"], documento, data=data, motorista=motorista, cliente=cliente, delete_incomplete_only=incomplete_only)
    for ws in wb.worksheets:
        if ws.title in ["Mestre", "Regras"]:
            continue
        total += delete_record_from_ws(ws, documento, data=data, motorista=motorista, cliente=cliente, delete_incomplete_only=incomplete_only)
    add_rules_sheet(wb)
    style_all(wb)
    wb.save(MASTER_FILE)
    return total

st.set_page_config(page_title=APP, layout="wide")
st.title(APP)

st.info("Sistema gratuito/local/online: importa Excel, guarda histórico, permite busca, edição e exclusão manual.")

with st.sidebar:
    st.header("Arquivo mestre")
    if not MASTER_FILE.exists():
        if st.button("Criar Excel mestre vazio"):
            create_empty_workbook()
            st.success("Criado.")
    st.write(f"Arquivo local: `{MASTER_FILE}`")
    if MASTER_FILE.exists():
        with open(MASTER_FILE, "rb") as f:
            st.download_button("Baixar Excel mestre", f, file_name="excel_mestre_operacional.xlsx")

tab1, tab2, tab3, tab4 = st.tabs(["Importar Excel atualizado", "Buscar / Editar / Excluir", "Adicionar manual", "Histórico"])

with tab1:
    st.subheader("Importar Excel atualizado")
    st.write("Use aqui o Excel que o ChatGPT te enviou. O sistema consolida no banco local/online sem apagar histórico.")
    uploaded = st.file_uploader("Enviar Excel atualizado (.xlsx)", type=["xlsx"])
    if uploaded and st.button("Importar e consolidar"):
        qtd = import_excel(uploaded)
        st.success(f"Importação concluída. Registros processados: {qtd}")

with tab2:
    st.subheader("Busca operacional com edição e exclusão")
    df = workbook_to_df()
    q = st.text_input("Pesquisar por delivery, final, motorista, cliente, CPF, placa, unidade, data, remessa...")
    if q:
        qn = norm_upper(q)
        mask = df.apply(lambda r: qn in " ".join([norm_upper(x) for x in r.values]), axis=1)
        result = df[mask].copy()
    else:
        result = df.copy()

    st.dataframe(result, use_container_width=True, hide_index=True)
    st.caption(f"{len(result)} registro(s) encontrado(s).")

    st.markdown("---")
    st.subheader("Editar um registro encontrado")

    if result.empty:
        st.warning("Nenhum registro encontrado para editar.")
    else:
        options = []
        for i, row in result.reset_index().iterrows():
            label = f"{i+1} | {norm(row.get('Data'))} | {norm(row.get('Motorista'))} | {norm(row.get('Documento'))} | {norm(row.get('Cliente'))}"
            options.append((label, row.to_dict()))
        labels = [x[0] for x in options]
        selected_label = st.selectbox("Escolha o registro", labels)
        selected = dict(options[labels.index(selected_label)][1])
        original_doc = norm(selected.get("Documento"))
        original_data = norm(selected.get("Data"))

        with st.form("editar_registro"):
            cols1 = st.columns(3)
            data = cols1[0].text_input("Data", value=norm(selected.get("Data")))
            motorista = cols1[1].text_input("Motorista", value=norm(selected.get("Motorista")))
            documento = cols1[2].text_input("Documento", value=norm(selected.get("Documento")))

            cols2 = st.columns(4)
            cliente = cols2[0].text_input("Cliente", value=norm(selected.get("Cliente")))
            unidade = cols2[1].text_input("Unidade", value=norm(selected.get("Unidade")))
            paletes = cols2[2].text_input("Paletes", value=norm(selected.get("Paletes")))
            valor = cols2[3].text_input("Valor", value=norm(selected.get("Valor")))

            cols3 = st.columns(6)
            L = cols3[0].text_input("L", value=norm(selected.get("L")))
            C = cols3[1].text_input("C", value=norm(selected.get("C")))
            F = cols3[2].text_input("F", value=norm(selected.get("F")))
            status = cols3[3].text_input("Status", value=norm(selected.get("Status")))
            tipo = cols3[4].text_input("Tipo", value=norm(selected.get("Tipo")))
            cpf = cols3[5].text_input("CPF", value=norm(selected.get("CPF")))

            cols4 = st.columns(6)
            cavalo = cols4[0].text_input("Cavalo", value=norm(selected.get("Cavalo")))
            carreta = cols4[1].text_input("Carreta", value=norm(selected.get("Carreta")))
            CS_OK = cols4[2].checkbox("CS OK", value=bool(selected.get("CS_OK")))
            C_OK = cols4[3].checkbox("C OK", value=bool(selected.get("C_OK")))
            L_OK = cols4[4].checkbox("L OK", value=bool(selected.get("L_OK")))
            F_OK = cols4[5].checkbox("F OK", value=bool(selected.get("F_OK")))

            obs = st.text_area("Observações", value=norm(selected.get("Observações")))
            inc = st.text_area("Inconsistências", value=norm(selected.get("Inconsistências")))

            salvar = st.form_submit_button("Salvar alteração")
            if salvar:
                new_row = {
                    "Data": data, "Motorista": motorista, "Documento": documento,
                    "Cliente": cliente, "Unidade": unidade, "Paletes": paletes, "Valor": valor,
                    "L": L, "C": C, "F": F, "Status": status, "Tipo": tipo,
                    "CPF": cpf, "Cavalo": cavalo, "Carreta": carreta,
                    "CS_OK": CS_OK, "C_OK": C_OK, "L_OK": L_OK, "F_OK": F_OK,
                    "Observações": obs, "Inconsistências": inc
                }
                edit_record(original_doc, original_data, new_row)
                st.success("Registro alterado. Refaça a busca ou atualize a página.")

        st.markdown("---")
        st.subheader("Excluir registros")
        st.warning("A exclusão cria backup antes de apagar.")

        colx1, colx2 = st.columns(2)
        with colx1:
            if st.button("Excluir SOMENTE o registro escolhido"):
                qtd = delete_record(original_doc, data=original_data, motorista=selected.get("Motorista"), cliente=selected.get("Cliente"), incomplete_only=False)
                st.success(f"Registro(s) excluído(s): {qtd}. Refaça a busca.")
        with colx2:
            if st.button("Excluir apenas linhas incompletas dessa delivery"):
                qtd = delete_record(original_doc, incomplete_only=True)
                st.success(f"Linha(s) incompleta(s) excluída(s): {qtd}. Refaça a busca.")

with tab3:
    st.subheader("Adicionar ou corrigir registro manualmente")
    with st.form("manual"):
        c1, c2, c3 = st.columns(3)
        data = c1.text_input("Data", placeholder="15/05/2026")
        motorista = c2.text_input("Motorista")
        documento = c3.text_input("Delivery / Remessa")

        c4, c5, c6 = st.columns(3)
        cliente = c4.text_input("Cliente")
        unidade = c5.text_input("Unidade")
        tipo = c6.selectbox("Tipo", ["delivery", "remessa", "deslocamento", "sem coleta", "não identificado"])

        c7, c8, c9, c10, c11 = st.columns(5)
        paletes = c7.text_input("Paletes")
        valor = c8.text_input("Valor")
        L = c9.text_input("L")
        C = c10.text_input("C")
        F = c11.text_input("F")

        c12, c13, c14, c15 = st.columns(4)
        CS_OK = c12.checkbox("CS OK")
        C_OK = c13.checkbox("C OK")
        L_OK = c14.checkbox("L OK")
        F_OK = c15.checkbox("F OK")

        status = st.text_input("Status")
        obs = st.text_area("Observações")
        inc = st.text_area("Inconsistências")

        submitted = st.form_submit_button("Salvar no Excel mestre")
        if submitted:
            try:
                row = {
                    "Data": data, "Motorista": motorista, "Documento": documento,
                    "Cliente": cliente, "Unidade": unidade, "Paletes": paletes, "Valor": valor,
                    "L": L, "C": C, "F": F, "Status": status, "Tipo": tipo,
                    "CS_OK": CS_OK, "C_OK": C_OK, "L_OK": L_OK, "F_OK": F_OK,
                    "Observações": obs, "Inconsistências": inc
                }
                add_manual_record(row)
                st.success("Registro salvo/atualizado.")
            except Exception as e:
                st.error(str(e))

with tab4:
    st.subheader("Histórico")
    df = workbook_to_df()
    if df.empty:
        st.warning("Ainda não há registros.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Registros", len(df))
        c2.metric("Motoristas", df["Motorista"].dropna().nunique())
        c3.metric("Clientes", df["Cliente"].dropna().nunique())
        c4.metric("Dias", df["Data"].dropna().nunique())

        st.write("Últimos registros:")
        st.dataframe(df.tail(50), use_container_width=True, hide_index=True)
